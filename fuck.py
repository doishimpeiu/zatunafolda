from openpyxl import Workbook
wb = Workbook()
ws = wb.active

a = []
b = []
c = []
d = []
e = []
f = []
g = []
h = []
i = []
j = []
k = []

# a.append(0)
for l in range(227):
  # print(f"{l}行目")
  for x in range(21):
    a.append(x * 11 + 227 * l)
    b.append(x * 11 + 1 + 227 * l)
    c.append(x * 11 + 2 + 227 * l)
    d.append(x * 11 + 3 + 227 * l)
    e.append(x * 11 + 4 + 227 * l)
    f.append(x * 11 + 5 + 227 * l)
    g.append(x * 11 + 6 + 227 * l)
  for x in range(20):
    h.append(x * 11 + 7 + 227 * l)
    i.append(x * 11 + 8 + 227 * l)
    j.append(x * 11 + 9 + 227 * l)
    k.append(x * 11 + 10 + 227 * l)

for n in range(4767):
  ws.cell(1+ n, 1, value=a[n])
  ws.cell(1+ n, 2, value=b[n])
  ws.cell(1+ n, 3, value=c[n])
  ws.cell(1+ n, 4, value=d[n])
  ws.cell(1+ n, 5, value=e[n])
  ws.cell(1+ n, 6, value=f[n])
for n in range(4540):
  ws.cell(1+ n, 7, value=g[n])
  ws.cell(1+ n, 8, value=h[n])
  ws.cell(1+ n, 9, value=i[n])
  ws.cell(1+ n, 10, value=j[n])
  ws.cell(1+ n, 11, value=k[n])


wb.save('fuck.xlsx')